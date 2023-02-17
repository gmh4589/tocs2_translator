# -*- coding:utf-8 -*-
import os
import shutil
import webbrowser

import openpyxl
import configparser

from datetime import datetime

import pyperclip
from kivy.app import App
from kivy.core.window import Window
from kivy.factory import Factory
from kivy.lang import Builder
from kivy.clock import Clock
from kivy.uix.screenmanager import ScreenManager, Screen, CardTransition

from tkinter import filedialog as fd

from kivymd.app import MDApp
from kivymd.uix.button import MDFlatButton
from kivymd.uix.dialog import MDDialog

setting = configparser.ConfigParser()
setting.read('setting.ini')

autosave = int(setting['Main']['autosave'])
backup = int(setting['Main']['backup'])
translator = setting['Main']['translator']

fs = int(setting['Size']['font'])

try: lastPath = setting['File']['path']
except KeyError: lastPath = os.environ['USERPROFILE'] + '/Desktop'

# Загружает конфиг интерфейса
Builder.load_file('new_translator.kv')
Builder.load_file('setting.kv')

class Colors:
    r = float(setting['Color']['r']) if float(setting['Color']['r']) <= 1 else 1/255 * int(setting['Color']['r'])
    g = float(setting['Color']['g']) if float(setting['Color']['g']) <= 1 else 1/255 * int(setting['Color']['g'])
    b = float(setting['Color']['b']) if float(setting['Color']['b']) <= 1 else 1/255 * int(setting['Color']['b'])
    a = float(setting['Color']['a']) if float(setting['Color']['a']) <= 1 else 1/255 * int(setting['Color']['a'])

    inputSize = float(setting['Size']['text'])
    lastFile = str(setting['File']['lastFile'])
    position = int(setting['File']['position'])
    allStrings = int(setting['File']['allStrings'])
    fontSize = fs
    print(r, g, b, a)


class MainScreen(Screen, Colors):

    values = []
    filename = ''
    sheet = ''
    readData = ''
    sss = 0

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        Clock.schedule_interval(self.update, .1)
        Clock.schedule_interval(self.autoSave, autosave)
        Clock.schedule_interval(self.backups, backup)

        if self.allStrings > self.position:
            self.fileOpen(self.lastFile)

    def fileOpen(self, filename = ''):
        if not self.values:
            if filename == '':
                filetypes = (('Файлы Excel', '*.xlsx'),
                            ('All files', '*.*'))

                self.filename = fd.askopenfilename(title = 'Выберите XLSX файл',
                                                    initialdir = lastPath,
                                                    filetypes = filetypes)
            else: self.filename = filename

            if self.filename == self.lastFile and self.allStrings == self.position and self.filename != '':
                Factory.MessageBox(title = 'Файл уже переведен! Откройте другой файл.').open()

            elif self.filename != '':

                print(self.filename)
                fName = self.filename.split('/')[-1]

                self.readData = openpyxl.load_workbook(self.filename, data_only = True)
                self.sheet = self.readData.active
                rows = self.sheet.max_row
                cols = self.sheet.max_column

                p = 0
                for row in range(1, rows):
                    percent = int((100/rows) * row)
                    if percent != p and percent % 10 == 0: print(f'ЗАГРУЗКА {percent}%...')
                    p = percent
                    for col in range(1, cols):
                        cell = self.sheet.cell(row = row, column = col).value

                        if cell == 'dialog':
                            val = [self.sheet.cell(row = row + 1, column = col).value, '', row + 1, col]
                            self.values.append(val)

                setting.set('File', 'path', self.filename.replace(fName, ''))
                setting.set('File', 'lastFile', self.filename)
                setting.set('File', 'allStrings', str(len(self.values)))

                with open('setting.ini', "w") as config_file:
                    setting.write(config_file)

                if self.values:
                    print('ФАЙЛ ЗАГРУЖЕН!')
                    if self.filename != self.lastFile:
                        self.nextString('0')
                    else: self.nextString(str(self.position))
                else: Factory.MessageBox(title = 'В файле не найдено текста для перевода!').open()

                #print(self.values)
        else: Factory.MessageBox(title = 'Сначала закройте открытый файл!').open()

    def writeFile(self, auto = 0):

        print(auto)

        if self.filename != '':
            for cell in range(len(self.values)):
                o = 1 if self.values[cell][1] != '' else 0
                self.sheet.cell(row = int(self.values[cell][2]),
                                column = int(self.values[cell][3])).value = self.values[cell][o]

            self.readData.save(self.filename)
            now = datetime.now().strftime('%d.%m.%Y %H-%M-%S')
            if auto == 0:
                #Factory.MessageBox(title = 'Перевод сохранен!').open()
                MDDialog(title = 'СООБЩЕНИЕ', text = 'Перевод сохранен!', buttons = [MDFlatButton(text = "OK", on_press = self.close())]).open()
            else:
                if autosave != 0: print(f'{now} - Автосохренение выполнено!')

    def createBackup(self):

        if self.filename != '':
            now = datetime.now().strftime('%d.%m.%Y %H-%M-%S')
            name = self.filename.split('/')[-1].split('.')[0]
            shutil.copyfile(self.filename, f'backups/{name}.{now}.xlsx')
            print(f'{now} - Бекап создан!')

    def autoSave(self, dt):
        if self.filename != '' and autosave != 0: self.writeFile(1)

    def backups(self, dt):
        if self.filename != '' and backup != 0: self.createBackup()

    def deleteText(self):
        self.ids['text4find'].text = ''
        self.ids['text4replace'].text = ''

    def copyText(self, where):
        if where == 'tl': pyperclip.copy(self.ids['newTXT'].text)
        if where == 'or': pyperclip.copy(self.ids['originalTXT'].text)

    def closeFile(self):
        if self.filename != '':
            self.values = []
            self.ids['originalTXT'].text = ''
            self.ids['newTXT'].text = ''
            self.ids['rdyLabel'].text = ''
            self.ids['longLabel'].text = ''
            self.ids['percentLabel'].text = ''
            self.ids['progressBar'].value = 0
            self.sss = 0

            setting.set('File', 'lastFile', '')
            setting.set('File', 'allStrings', '0')
            setting.set('File', 'position', '0')

            with open('setting.ini', "w") as config_file:
                setting.write(config_file)

            Factory.MessageBox(title = f'Файл {self.filename} закрыт!', size = (300, 150)).open()

    def nextString(self, step = '+'):

        def getHead(t):
            i = 0

            t = t.replace("'", '')

            for i in range(len(t)):
                if t[i] == '.':
                    i += 1
                    break
                if t[i].isalpha():
                    if t[i].islower(): break
            else:
                if t[-1] == 'I' or t[-1] == 'A': i = 0

            n = (i - 1)
            if "(" in t: n -= 1
            return n

        if len(self.values) != 0:
            try:
                old = self.values[self.sss][0].split(' ')[0]
                h = getHead(old.split(' ')[0])
                head = old.split(' ')[0][:h]
                self.values[self.sss][1] = head + self.ids['newTXT'].text
                print(self.values[self.sss][1])

                if step == '+': self.sss += 1
                elif step == '-':
                    if self.sss != 0: self.sss -= 1
                    #else: self.EOF()
                else: self.sss = int(step)

                o = 1 if self.values[self.sss][1] != '' else 0

                ttt = str(self.values[self.sss][o])

                hText = ttt.split(' ')[0]
                if hText[0] == '#': head = getHead(hText)
                else: head = 0
                self.ids['newTXT'].text = ttt.replace(hText[:head], '')
                self.ids['originalTXT'].text = ttt

                setting.set('File', 'position', str(self.sss))
                with open('setting.ini', "w") as config_file:
                    setting.write(config_file)

            except IndexError:
                self.writeFile(1)
                Factory.MessageBox(title = 'Конец файла! Перевод сохранен!').open()
                setting.set('File', 'position', str(self.allStrings))

                with open('setting.ini', "w") as config_file:
                    setting.write(config_file)

    def update(self, dt):

        global fs

        self.ids['upPanel'].size_hint = (1, 1/Window.size[1] * 70)
        self.ids['originalTXT'].font_size = fs
        self.ids['newTXT'].font_size = fs

        try:
            self.ids['originalTXT'].text = str(self.values[self.sss][0])
            self.ids['rdyLabel'].text = str(self.sss + 1) + '/' + str(len(self.values)) + ' готово'
            self.ids['longLabel'].text = str(len(self.values[self.sss][0])) + ' символов'
            percent = round(100 / len(self.values) * (self.sss + 1), 2)
            self.ids['percentLabel'].text = str(percent) + ' %'
            self.ids['progressBar'].value = percent

        except IndexError: pass

        s = configparser.ConfigParser()
        s.read('setting.ini')

        iS = float(s['Size']['text'])

        if iS != self.inputSize:
            self.inputSize = iS
            self.ids['text4find'].size_hint = (1, iS/10)
            self.ids['text4replace'].size_hint = (1, iS/10)
            self.ids['originalTXT'].size_hint = (1, .9 - iS/10)
            self.ids['newTXT'].size_hint = (1, .9 - iS/10)

    def reFresh(self):
        self.ids['newTXT'].text = self.ids['originalTXT'].text

    def findBTN(self):

        text4find = str(self.ids['text4find'].text)
        text4replace = str(self.ids['text4replace'].text)
        count = 0

        if text4find != '' and text4replace != '':

            for st in range(len(self.values)):
                o = 1 if self.values[st][1] != '' else 0
                if self.values[st][o].find(text4find) != -1:
                    self.values[st][1] = self.values[st][o].replace(text4find, text4replace)
                    if st == self.sss: self.upd()
                    count += 1

            Factory.MessageBox(title = f'Выполнено {str(count)} замен').open()

    def upd(self):
        if str(self.values[self.sss][1]) == '':
            self.ids['newTXT'].text = str(self.values[self.sss][0])
        else:
            self.ids['newTXT'].text = str(self.values[self.sss][1])

    def gotoBTN(self):
        if len(self.values) != 0:
            try:
                goto = int(self.ids['gotoInput'].text)
                if goto < len(self.values) + 1: self.nextString(str(goto - 1))
                else: Factory.MessageBox(title = f'Вы ввели {goto}, а в файле только {len(self.values)} строк!').open()
            except (TypeError, ValueError):
                Factory.MessageBox(title = f'Введите число от 1 до {len(self.values)}!').open()

    def translate(self):

        if setting['Main']['translator'] == 'Yandex': site = 'https://translate.yandex.ru/?utm_source=main_stripe_big&lang=en-ru&text='
        elif setting['Main']['translator'] == 'Google': site = 'https://translate.google.ru/?hl=ru&tab=rT&sl=en&tl=ru&text='
        elif setting['Main']['translator'] == 'Bing': site = 'https://www.bing.com/translator/?text='
        elif setting['Main']['translator'] == 'DeepL': site = 'https://www.deepl.com/translator#en/ru/'
        elif setting['Main']['translator'] == 'Promt': site = 'https://www.translate.ru/%D0%BF%D0%B5%D1%80%D0%B5%D0%B2%D0%BE%D0%B4/%D0%B0%D0%BD%D0%B3%D0%BB%D0%B8%D0%B9%D1%81%D0%BA%D0%B8%D0%B9-%D1%80%D1%83%D1%81%D1%81%D0%BA%D0%B8%D0%B9?text='
        else: site = ''

        t = self.ids['newTXT'].text.replace('#', '')
        if t != '' and site != '': webbrowser.open(site + t)

class SettingScreen(Screen, Colors):

    tl = ''

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.ids['autoSavePeriod'].text = str(autosave)
        self.ids['backupPeriod'].text = str(backup)

        self.ids[translator].state = 'down'

        Clock.schedule_interval(self.update, .1)

    def checkboxClick(self, tl1): self.tl = tl1

    def setSetting(self):

        setting.set('Main', 'autosave', self.ids['autoSavePeriod'].text)
        setting.set('Main', 'backup', self.ids['backupPeriod'].text)
        setting.set('Main', 'translator', self.tl)

        setting.set('Color', 'a', self.ids['alphaInfo'].text)
        setting.set('Size', 'text', self.ids['sizeInfo'].text)
        setting.set('Size', 'font', self.ids['fontSizeInfo'].text)

        with open('setting.ini', "w") as config_file:
            setting.write(config_file)

        self.manager.current = 'main'

    def update(self, dt):
        global fs
        self.ids['sizeInfo'].text = str(round(self.ids['sizeData'].value, 1))
        self.ids['alphaInfo'].text = str(round(self.ids['alphaData'].value, 2))
        fs = int(self.ids['fontSize'].value)
        self.ids['fontSizeInfo'].text = str(fs)
        self.fontSize = fs


class NewTranslatorApp(MDApp):

    # Создаёт интерфейс
    def build(self):
        self.theme_cls.theme_style = "Dark"
        self.theme_cls.primary_palette = "Green"
        sm = ScreenManager(transition = CardTransition())
        sm.add_widget(MainScreen(name = 'main'))
        sm.add_widget(SettingScreen(name = 'setting'))

        return sm

NewTranslatorApp().run()
